from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = 'Quote'

# Page setup
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.orientation = 'portrait'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.page_margins.left = 0.5
ws.page_margins.right = 0.5
ws.page_margins.top = 0.4
ws.page_margins.bottom = 0.3

# Hide gridlines
ws.sheet_view.showGridLines = False

# 5 columns for flexible layout: A=indent, B=labels, C=values, D=spacer, E=amounts
ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 42
ws.column_dimensions['D'].width = 4
ws.column_dimensions['E'].width = 16

# Colours
navy = '1B2A4A'
teal = '00B4D8'
charcoal = '2C3E50'
grey_label = '4A5568'
grey_desc = '718096'
grey_footer = 'A0AEC0'
white = 'FFFFFF'
light_bg = 'F5F6FA'
teal_light = 'E8F9FD'
border_light = 'E2E8F0'

# Google brand colours (from ACT logo)
g_blue = '4285F4'
g_red = 'EA4335'
g_yellow = 'FBBC05'
g_green = '34A853'

# Fills
fill_navy = PatternFill('solid', fgColor=navy)
fill_white = PatternFill('solid', fgColor=white)
fill_light = PatternFill('solid', fgColor=light_bg)
fill_teal_light = PatternFill('solid', fgColor=teal_light)
fill_teal = PatternFill('solid', fgColor=teal)

# Borders
no_border = Border(
    left=Side(style=None), right=Side(style=None),
    top=Side(style=None), bottom=Side(style=None)
)
border_bottom_navy = Border(bottom=Side(style='medium', color=navy))
border_bottom_teal = Border(bottom=Side(style='medium', color=teal))
border_bottom_light = Border(bottom=Side(style='thin', color=border_light))
border_top_navy = Border(top=Side(style='medium', color=navy))

def fill_row_white(row, cols=5):
    for c in range(1, cols + 1):
        ws.cell(row, c).fill = fill_white

def fill_row_bg(row, fill, cols=5):
    for c in range(1, cols + 1):
        ws.cell(row, c).fill = fill

def merge_content(row, start_col, end_col):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)

r = 1

# === HEADER BAR — navy background ===
fill_row_bg(r, fill_navy)
ws.row_dimensions[r].height = 10
r += 1

# Brand accent stripe — 4 thin coloured cells
fill_row_bg(r, fill_white)
ws.row_dimensions[r].height = 4
ws.cell(r, 1).fill = PatternFill('solid', fgColor=g_blue)
ws.cell(r, 2).fill = PatternFill('solid', fgColor=g_red)
ws.cell(r, 3).fill = PatternFill('solid', fgColor=g_yellow)
ws.cell(r, 4).fill = PatternFill('solid', fgColor=g_green)
ws.cell(r, 5).fill = PatternFill('solid', fgColor=g_blue)
r += 1

# Spacer
fill_row_white(r)
ws.row_dimensions[r].height = 12
r += 1

# Name row with logo
fill_row_white(r)
logo_row = r
merge_content(r, 2, 3)
ws.cell(r, 2, 'Christopher Hoole').font = Font(name='Arial', size=20, bold=True, color=navy)
ws.cell(r, 2).alignment = Alignment(horizontal='left', vertical='center')
merge_content(r, 4, 5)
ws.cell(r, 4, 'christopherhoole.com').font = Font(name='Arial', size=9, color=grey_label)
ws.cell(r, 4).alignment = Alignment(horizontal='right', vertical='center')
ws.row_dimensions[r].height = 30
r += 1

# Subtitle + contact
fill_row_white(r)
merge_content(r, 2, 3)
ws.cell(r, 2, 'Google Ads Specialist').font = Font(name='Arial', size=11, bold=True, color=teal)
ws.cell(r, 2).alignment = Alignment(horizontal='left', vertical='center')
merge_content(r, 4, 5)
ws.cell(r, 4, 'chris@christopherhoole.com  |  +44 7XXX XXX XXX').font = Font(name='Arial', size=8, color=grey_label)
ws.cell(r, 4).alignment = Alignment(horizontal='right', vertical='center')
ws.row_dimensions[r].height = 18
r += 1

# Navy divider line
fill_row_white(r)
for c in range(1, 6):
    ws.cell(r, c).border = border_bottom_navy
ws.row_dimensions[r].height = 4
r += 1

# Spacer
fill_row_white(r)
ws.row_dimensions[r].height = 14
r += 1

# === QUOTE title ===
fill_row_white(r)
merge_content(r, 1, 5)
ws.cell(r, 1, 'QUOTE').font = Font(name='Arial', size=16, bold=True, color=navy)
ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[r].height = 28
r += 1

# Spacer
fill_row_white(r)
ws.row_dimensions[r].height = 6
r += 1

# === QUOTE FOR — light background block ===
fill_row_bg(r, fill_light)
merge_content(r, 1, 5)
ws.cell(r, 1, '  QUOTE FOR').font = Font(name='Arial', size=9, bold=True, color=navy)
ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[r].height = 22
r += 1

meta = [
    ('Client:', 'Owen Hoare', 'Date:', '30 March 2026'),
    ('Company:', 'Objection Experts', 'Quote Ref:', 'CH-OE-001'),
    ('Website:', 'objectionexperts.com', 'Valid Until:', '30 April 2026'),
]
for label1, val1, label2, val2 in meta:
    fill_row_white(r)
    ws.cell(r, 2, label1).font = Font(name='Arial', size=9, bold=True, color=grey_label)
    ws.cell(r, 2).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(r, 3, val1).font = Font(name='Arial', size=9, color=charcoal)
    ws.cell(r, 3).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(r, 4, label2).font = Font(name='Arial', size=9, bold=True, color=grey_label)
    ws.cell(r, 4).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(r, 5, val2).font = Font(name='Arial', size=9, color=charcoal)
    ws.cell(r, 5).alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[r].height = 18
    r += 1

# Spacer
fill_row_white(r)
ws.row_dimensions[r].height = 10
r += 1

# === SCOPE OF WORK ===
fill_row_bg(r, fill_light)
merge_content(r, 1, 5)
ws.cell(r, 1, '  SCOPE OF WORK').font = Font(name='Arial', size=9, bold=True, color=navy)
ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[r].height = 22
r += 1

scope_items = [
    'Google Ads account audit and restructure (as per Reports 1\u20133)',
    'Conversion tracking fix \u2014 Email Click reclassified to Secondary',
    'Bid strategy optimisation with target CPA implementation',
    'Negative keyword framework implementation',
    'Ad copy A/B testing (3 RSA variants per ad group)',
    'Ad schedule and device bid adjustments',
    'Weekly search term review and negative keyword management',
    'Monthly performance report',
    'Ongoing campaign optimisation and strategic management',
]
for item in scope_items:
    fill_row_white(r)
    ws.cell(r, 2, '\u2022').font = Font(name='Arial', size=9, color=teal, bold=True)
    ws.cell(r, 2).alignment = Alignment(horizontal='right', vertical='center')
    merge_content(r, 3, 5)
    ws.cell(r, 3, item).font = Font(name='Arial', size=9, color=charcoal)
    ws.cell(r, 3).alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 17
    r += 1

# Spacer
fill_row_white(r)
ws.row_dimensions[r].height = 10
r += 1

# === PRICING ===
fill_row_bg(r, fill_light)
merge_content(r, 1, 5)
ws.cell(r, 1, '  PRICING').font = Font(name='Arial', size=9, bold=True, color=navy)
ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[r].height = 22
r += 1

# Table header
fill_row_bg(r, fill_navy)
merge_content(r, 1, 4)
ws.cell(r, 1, '  ITEM').font = Font(name='Arial', size=8.5, bold=True, color=white)
ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
ws.cell(r, 5, 'AMOUNT').font = Font(name='Arial', size=8.5, bold=True, color=white)
ws.cell(r, 5).alignment = Alignment(horizontal='right', vertical='center')
ws.row_dimensions[r].height = 24
r += 1

# Setup fee
fill_row_white(r)
merge_content(r, 1, 4)
ws.cell(r, 1, '  Setup Fee').font = Font(name='Arial', size=9.5, bold=True, color=charcoal)
ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
ws.cell(r, 5, '\u00a3[SETUP_FEE]').font = Font(name='Arial', size=9.5, bold=True, color=charcoal)
ws.cell(r, 5).alignment = Alignment(horizontal='right', vertical='center')
ws.row_dimensions[r].height = 20
r += 1

fill_row_white(r)
merge_content(r, 1, 5)
ws.cell(r, 1, '  One-off \u2014 account restructure and implementation of all quick wins from the audit reports').font = Font(name='Arial', size=8, italic=True, color=grey_desc)
ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
for c in range(1, 6):
    ws.cell(r, c).border = border_bottom_light
ws.row_dimensions[r].height = 15
r += 1

# Monthly fee
fill_row_white(r)
merge_content(r, 1, 4)
ws.cell(r, 1, '  Monthly Management Fee').font = Font(name='Arial', size=9.5, bold=True, color=charcoal)
ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
ws.cell(r, 5, '\u00a3[MONTHLY_FEE] /mo').font = Font(name='Arial', size=9.5, bold=True, color=charcoal)
ws.cell(r, 5).alignment = Alignment(horizontal='right', vertical='center')
ws.row_dimensions[r].height = 20
r += 1

fill_row_white(r)
merge_content(r, 1, 5)
ws.cell(r, 1, '  Ongoing optimisation, monitoring, reporting, and strategic management').font = Font(name='Arial', size=8, italic=True, color=grey_desc)
ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
for c in range(1, 6):
    ws.cell(r, c).border = border_bottom_light
ws.row_dimensions[r].height = 15
r += 1

# Total row — teal highlight
fill_row_bg(r, fill_teal_light)
merge_content(r, 1, 4)
ws.cell(r, 1, '  Total Due on Acceptance').font = Font(name='Arial', size=10.5, bold=True, color=navy)
ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
ws.cell(r, 5, '\u00a3[SETUP_FEE]').font = Font(name='Arial', size=10.5, bold=True, color=navy)
ws.cell(r, 5).alignment = Alignment(horizontal='right', vertical='center')
ws.row_dimensions[r].height = 26
r += 1

# Spacer
fill_row_white(r)
ws.row_dimensions[r].height = 10
r += 1

# === WHAT'S INCLUDED ===
fill_row_bg(r, fill_light)
merge_content(r, 1, 5)
ws.cell(r, 1, "  WHAT'S INCLUDED").font = Font(name='Arial', size=9, bold=True, color=navy)
ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[r].height = 22
r += 1

included = [
    ('Weekly campaign optimisation', 'Monthly performance report'),
    ('Search term reviews & negatives', 'Ad copy A/B testing & refresh'),
    ('Bid strategy management', 'Conversion tracking monitoring'),
    ('Ad schedule optimisation', 'Email support'),
]
for a, b in included:
    fill_row_white(r)
    merge_content(r, 1, 3)
    ws.cell(r, 1, '  \u2713  ' + a).font = Font(name='Arial', size=8.5, color=charcoal)
    ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
    merge_content(r, 4, 5)
    ws.cell(r, 4, '\u2713  ' + b).font = Font(name='Arial', size=8.5, color=charcoal)
    ws.cell(r, 4).alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 17
    r += 1

# Spacer
fill_row_white(r)
ws.row_dimensions[r].height = 10
r += 1

# === PAYMENT TERMS ===
fill_row_bg(r, fill_light)
merge_content(r, 1, 5)
ws.cell(r, 1, '  PAYMENT TERMS').font = Font(name='Arial', size=9, bold=True, color=navy)
ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[r].height = 22
r += 1

payments = [
    'Setup fee due on acceptance of this quote',
    'Monthly fee billed on the 1st of each month, starting the month after setup',
    'Payment by bank transfer: Christopher Hoole, Acc: 04345312, Sort: 04-00-75',
    '30-day notice period to cancel',
]
for p in payments:
    fill_row_white(r)
    ws.cell(r, 2, '\u2022').font = Font(name='Arial', size=9, color=grey_label)
    ws.cell(r, 2).alignment = Alignment(horizontal='right', vertical='center')
    merge_content(r, 3, 5)
    ws.cell(r, 3, p).font = Font(name='Arial', size=8.5, color=charcoal)
    ws.cell(r, 3).alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 17
    r += 1

# Spacer
fill_row_white(r)
ws.row_dimensions[r].height = 10
r += 1

# === TERMS & CONDITIONS ===
fill_row_bg(r, fill_light)
merge_content(r, 1, 5)
ws.cell(r, 1, '  TERMS & CONDITIONS').font = Font(name='Arial', size=9, bold=True, color=navy)
ws.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[r].height = 22
r += 1

terms = [
    'Minimum contract period: 3 months after setup completion (to allow optimisations to take effect)',
    '30-day written notice required to terminate after the minimum period',
    'Ad spend is billed directly by Google and is not included in the fees above',
    'Landing page design, website development, and creative assets are not included unless explicitly stated',
    'No specific performance outcomes are guaranteed. Results are influenced by budget, landing page quality, and market conditions.',
]
for t in terms:
    fill_row_white(r)
    ws.cell(r, 2, '\u2022').font = Font(name='Arial', size=9, color=grey_label)
    ws.cell(r, 2).alignment = Alignment(horizontal='right', vertical='center')
    merge_content(r, 3, 5)
    ws.cell(r, 3, t).font = Font(name='Arial', size=8.5, color=charcoal)
    ws.cell(r, 3).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 17
    r += 1

# Spacer
fill_row_white(r)
ws.row_dimensions[r].height = 14
r += 1

# === FOOTER ===
# Teal divider
fill_row_white(r)
for c in range(1, 6):
    ws.cell(r, c).border = Border(bottom=Side(style='thin', color=teal))
ws.row_dimensions[r].height = 2
r += 1

fill_row_white(r)
ws.row_dimensions[r].height = 6
r += 1

fill_row_white(r)
merge_content(r, 1, 5)
ws.cell(r, 1, 'Christopher Hoole  \u00b7  Google Ads Specialist  \u00b7  christopherhoole.com  \u00b7  chris@christopherhoole.com').font = Font(name='Arial', size=7.5, color=grey_footer)
ws.cell(r, 1).alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[r].height = 14
r += 1

fill_row_white(r)
merge_content(r, 1, 5)
ws.cell(r, 1, 'This quote is confidential and intended solely for the named recipient.').font = Font(name='Arial', size=7, italic=True, color=grey_footer)
ws.cell(r, 1).alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[r].height = 12
r += 1

# Brand stripe at bottom too
fill_row_bg(r, fill_white)
ws.row_dimensions[r].height = 4
ws.cell(r, 1).fill = PatternFill('solid', fgColor=g_blue)
ws.cell(r, 2).fill = PatternFill('solid', fgColor=g_red)
ws.cell(r, 3).fill = PatternFill('solid', fgColor=g_yellow)
ws.cell(r, 4).fill = PatternFill('solid', fgColor=g_green)
ws.cell(r, 5).fill = PatternFill('solid', fgColor=g_blue)
r += 1

# Fill remaining visible rows with white to hide gridlines
for extra in range(r, r + 20):
    fill_row_white(extra)

# Add logo image
logo = XlImage('potential_clients/objection_experts/act_logo.png')
logo.width = 38
logo.height = 38
ws.add_image(logo, f'A{logo_row}')

ws.print_area = f'A1:E{r - 1}'

wb.save('potential_clients/objection_experts/quote_objection_experts.xlsx')
print(f'Done. {r - 1} rows.')
